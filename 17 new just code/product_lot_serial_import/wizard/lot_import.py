# -*- coding: utf-8 -*-

import base64
import openpyxl
from io import BytesIO
import xlrd
import pytz
from datetime import datetime
from odoo import api, fields, models
from odoo.exceptions import ValidationError
from pytz import timezone
from odoo.tools import DEFAULT_SERVER_DATETIME_FORMAT

class ImportLotNumber(models.TransientModel):
    _name = 'import.serial.number'
    _description = "Import lot number"

    company_id = fields.Many2one(
        'res.company',
        string='Company',
        default=lambda self: self.env.user.company_id,
        required=True,
    )
    files = fields.Binary(
        string="Import Product by",
    )
    import_product_by = fields.Selection(
        [('name', 'Name'),
         ('code', 'Code'),
         ('barcode', 'Barcode')],
        string='Search Product by',
        default='name',
    )

    def _convert_to_utc(self, localdatetime):
        check_in_date = datetime.strptime(localdatetime, '%m/%d/%Y %H:%M:%S')
        timezone_tz = 'utc'
        user_id = self.env.user
        if user_id.tz:
            timezone_tz = user_id.tz
        local = pytz.timezone(timezone_tz)
        local_dt = local.localize(check_in_date, is_dst=None)
        utc_datetime = local_dt.astimezone(pytz.utc)
        return utc_datetime

    # @api.multi
    def do_import(self):
        try:
            workbook = xlrd.open_workbook(file_contents=base64.decodebytes(self.files))
        except:
            raise ValidationError("Please select .xls file...")
        
        product_lot_lists = []
        # production_lot_obj = self.env['stock.production.lot'] #odoo16
        production_lot_obj = self.env['stock.lot']
        sheet_name = workbook.sheet_names()
        sheet = workbook.sheet_by_name(sheet_name[0])
        number_of_rows = sheet.nrows
        sheet_number = sheet.row_values(0)
        serial_lot = sheet_number.index('Lot/Serial Number')
        product_value = sheet_number.index('Product')
        ref_var = sheet_number.index('Internal Reference')
        date = sheet_number.index('Best before Date')
        date_removal = sheet_number.index('Removal Date')
        date_end = sheet_number.index('End of Life Date')
        date_alert = sheet_number.index('Alert Date')
        row = 1
        while row < number_of_rows:
            lot_serial_number = sheet.cell(row, serial_lot).value
            if not lot_serial_number:
                raise ValidationError(
                    'Please add Lot / Serial Number %s' % (row+1)
                    )
            product = sheet.cell(row, product_value).value
            if not product:
                raise ValidationError('Please add product at %s.' % (row+1))
            ref = sheet.cell(row, ref_var).value
            before_date = sheet.cell(row, date).value
            # best_before_date = self._convert_to_utc(before_date).strftime\
            # ("%m/%d/%Y %H:%M:%S")
            best_before_date = self._convert_to_utc(before_date).strftime(DEFAULT_SERVER_DATETIME_FORMAT)
            date_remove = sheet.cell(row, date_removal).value
            # removal_date = self._convert_to_utc(date_remove).strftime\
            # ("%m/%d/%Y %H:%M:%S")
            removal_date = self._convert_to_utc(date_remove).strftime\
            (DEFAULT_SERVER_DATETIME_FORMAT)
            date_life = sheet.cell(row, date_end).value
            # end_of_life_date = self._convert_to_utc(date_life).strftime\
            # ("%m/%d/%Y %H:%M:%S")
            end_of_life_date = self._convert_to_utc(date_life).strftime\
            (DEFAULT_SERVER_DATETIME_FORMAT)
            date_alert_time = sheet.cell(row, date_alert).value
            # alert_date = self._convert_to_utc(date_alert_time).strftime\
            # ("%m/%d/%Y %H:%M:%S")
            alert_date = self._convert_to_utc(date_alert_time).strftime(DEFAULT_SERVER_DATETIME_FORMAT)
            if self.import_product_by == 'name' and product:
                product_id = self.env['product.product'].search(
                    [('name', '=', product)], limit=1
                )
            elif self.import_product_by == 'code' and product:
                product_id = self.env['product.product'].search(
                    [('default_code', '=', product)], limit=1
                )
            elif self.import_product_by == 'barcode' and product:
                product_id = self.env['product.product'].search(
                    [('barcode', '=', product)], limit=1
                )
            if not product_id:
                raise ValidationError(
                    'product should be empty at row number %s.' % (row+1)
                )
            vals = {
                'name': int(lot_serial_number),
                'product_id': product_id.id,
                'ref': ref,
                'use_date': best_before_date,
                'removal_date': removal_date,
                'expiration_date':end_of_life_date,
                # 'life_date': end_of_life_date,
                'alert_date': alert_date,
                'company_id':self.company_id.id,
            }
            lot_id = production_lot_obj.create(vals)
            product_lot_lists.append(lot_id.id)
            row += 1

        action = self.env.ref('stock.action_production_lot_form').sudo().read()[0]
        action['domain'] = [('id', 'in', product_lot_lists)]
        return action
